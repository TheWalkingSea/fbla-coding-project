import sys
import organization
import partner
import contacts
from openpyxl import Workbook
import inquirer
from utils import util

def report_menu() -> None:
    """ Represents the report menu that will give the user options on how they want their data exported. """
    util.cls()
    q = [
        inquirer.List("report", message="Choose an option to export the data", choices=[
            "Export organization",
            "Export partners",
            "Export organizations & partners",
            "Back"
        ])
    ]
    answers = inquirer.prompt(q)
    answer = answers['report']
    if (answer == "Export organization"):
        create_organization_report()
    elif (answer == "Export partners"):
        create_partner_report()
    elif (answer == "Export organizations & partners"):
        create_full_report()
    elif (answer == "Back"):
        from main import main
        main()

def create_partner_report() -> None:
    """ Creates a report for partners only """
    wb = Workbook()
    ws = wb.active
    orgdata = organization.get_all_data()
    partnerdata = partner.get_all_data()
    contactdata = contacts.get_all_data()

    def get_contact(id: int) -> dict:
        """ Gets contact data for a certain id using the contact data above 
        
        Parameters:
        (int)id: The id of the partner
        
        Returns:
        (dict): The data from the db
        
        """
        return list(filter(lambda x:x['id'] == id, contactdata))[0]
    
    ws.append(['Type', 'Name', 'Description', 'Email', 'Address', 'Phone', "Expertise", 'Role'])
    for org in orgdata:
        for ptnr in filter(lambda x: x['organization_fk'] == org['id'], partnerdata):
            partnercontact = get_contact(ptnr['contact_fk'])
            ws.append([
                'Partner', 
                partnercontact['name'], 
                ptnr['description'], 
                partnercontact['email'], 
                partnercontact['address'], 
                partnercontact['phone'], 
                ptnr['expertise'],
                ptnr['role'],
            ])

    try:
        wb.save("export.xlsx")
    except PermissionError:
        input("Something went wrong when trying to access the export file. Please delete the file and try again.")
        sys.exit(-1)

def create_organization_report() -> None:
    """ Creates a report for organizations only """
    wb = Workbook()
    ws = wb.active
    orgdata = organization.get_all_data()
    contactdata = contacts.get_all_data()

    def get_contact(id: int) -> dict:
        """ Gets contact data for a certain id using the contact data above 
        
        Parameters:
        (int)id: The id of the partner
        
        Returns:
        (dict): The data from the db
        
        """
        return list(filter(lambda x:x['id'] == id, contactdata))[0]
    
    ws.append(['Type', 'Name', 'Description', 'Email', 'Address', 'Phone', 'Type', 'URL'])
    for org in orgdata:
        orgcontact = get_contact(org['contact_fk'])
        ws.append([
            'Organization',
            orgcontact['name'], 
            org['description'], 
            orgcontact['email'], 
            orgcontact['address'], 
            orgcontact['phone'], 
            org['type'], 
            org['url']
        ])
    try:
        wb.save("export.xlsx")
    except PermissionError:
        input("Something went wrong when trying to access the export file. Please delete the file and try again.")
        sys.exit(-1)

def create_full_report() -> None:
    """ Creates a full Excel report and exports to export.xlsx """
    wb = Workbook()
    ws = wb.active
    orgdata = organization.get_all_data()
    partnerdata = partner.get_all_data()
    contactdata = contacts.get_all_data()

    def get_contact(id: int) -> dict:
        """ Gets contact data for a certain id using the contact data above 
        
        Parameters:
        (int)id: The id of the partner
        
        Returns:
        (dict): The data from the db
        
        """
        return list(filter(lambda x:x['id'] == id, contactdata))[0]
    
    ws.append(['Type', 'Name', 'Description', 'Email', 'Address', 'Phone', "Expertise", 'Role', 'Type', 'URL'])
    for org in orgdata:
        orgcontact = get_contact(org['contact_fk'])
        ws.append([
            'Organization',
            orgcontact['name'], 
            org['description'], 
            orgcontact['email'], 
            orgcontact['address'], 
            orgcontact['phone'], 
            '', 
            '', 
            org['type'], 
            org['url']
        ])
        for ptnr in filter(lambda x: x['organization_fk'] == org['id'], partnerdata):
            partnercontact = get_contact(ptnr['contact_fk'])
            ws.append([
                'Partner', 
                partnercontact['name'], 
                ptnr['description'], 
                partnercontact['email'], 
                partnercontact['address'], 
                partnercontact['phone'], 
                ptnr['expertise'],
                ptnr['role'],
                '',
                ''
            ])
        ws.append(['']*10)

    try:
        wb.save("export.xlsx")
    except PermissionError:
        input("Something went wrong when trying to access the export file. Please delete the file and try again.")
        sys.exit(-1)